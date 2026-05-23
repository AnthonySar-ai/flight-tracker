"""
generate_report.py
Reads data/flights.csv and data/no_results.csv and produces
data/flight_report.xlsx with multiple formatted sheets:

  1. Raw Data       — all collected offers, newest first
  2. Price Summary  — cheapest + average per depart_date per week
  3. Airline Stats  — average price and count per airline
  4. Disruption Log — dates that returned zero offers
  5. Best Deals     — top 20 cheapest offers ever seen
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

DATA_DIR    = Path("data")
CSV_PATH    = DATA_DIR / "flights.csv"
NO_RES_PATH = DATA_DIR / "no_results.csv"
OUT_PATH    = DATA_DIR / "flight_report.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
H_BG   = "1D6A96"   # header background (deep blue)
H_FG   = "FFFFFF"   # header text
AUG_BG = "E1F5EE"   # August row tint
DEC_BG = "E6F1FB"   # DecJan row tint
WARN   = "FFF3CD"   # warning yellow
GOOD   = "D4EDDA"   # good green
ALT    = "F8F9FA"   # alternating row

thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr_style(cell, text=None):
    if text:
        cell.value = text
    cell.font      = Font(bold=True, color=H_FG, name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=H_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border


def data_style(cell, align="left", bold=False, bg=None, num_fmt=None):
    cell.font      = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = border
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ── Load data ─────────────────────────────────────────────────────────────────

def load_data():
    if not CSV_PATH.exists():
        print(f"No data file found at {CSV_PATH}")
        return pd.DataFrame(), pd.DataFrame()

    df = pd.read_csv(CSV_PATH)
    df["price_eur"]     = pd.to_numeric(df["price_eur"], errors="coerce")
    df["checked_at"]    = pd.to_datetime(df["checked_at"], errors="coerce")
    df["depart_date"]   = pd.to_datetime(df["depart_date"], errors="coerce")
    df["airline_count"] = pd.to_numeric(df.get("airline_count", 0), errors="coerce").fillna(0).astype(int)
    df["price_change_pct"] = pd.to_numeric(df.get("price_change_pct", ""), errors="coerce")
    df = df.dropna(subset=["price_eur", "depart_date"])
    df = df.sort_values(["checked_at", "depart_date"], ascending=[False, True])

    no_res = pd.DataFrame()
    if NO_RES_PATH.exists():
        no_res = pd.read_csv(NO_RES_PATH)

    return df, no_res


# ── Sheet 1: Raw Data ─────────────────────────────────────────────────────────

def sheet_raw(wb, df):
    ws = wb.create_sheet("Raw Data")
    cols = ["checked_at","season","depart_date","return_date","price_eur",
            "price_change_pct","airlines","airline_count","out_stops",
            "out_dep_time","out_arr_time","out_duration"]
    labels = ["Checked","Season","Depart","Return","Price €",
              "Chg %","Airlines","# Airlines","Stops",
              "Dep time","Arr time","Duration (min)"]

    for i, lbl in enumerate(labels, 1):
        hdr_style(ws.cell(1, i), lbl)

    for r_idx, (_, row) in enumerate(df[cols].iterrows(), 2):
        bg = AUG_BG if str(row.get("season","")) == "August" else DEC_BG
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(r_idx, c_idx)
            val  = row[col]
            if pd.isna(val) or val == "":
                cell.value = ""
            elif col in ("checked_at","depart_date","return_date") and hasattr(val,"strftime"):
                cell.value = val.strftime("%Y-%m-%d")
            elif col == "price_eur":
                cell.value = round(float(val), 0)
                data_style(cell, "right", bold=True, bg=bg, num_fmt='€#,##0')
                continue
            elif col == "price_change_pct" and not pd.isna(val):
                cell.value = round(float(val), 1)
                cell.number_format = '+0.0%;-0.0%;0.0%'
                cell.font = Font(name="Arial", size=10,
                    color="A32D2D" if float(val) > 10 else ("3B6D11" if float(val) < -5 else "000000"))
                cell.alignment = Alignment(horizontal="right")
                cell.border = border
                cell.fill = PatternFill("solid", fgColor=bg)
                continue
            else:
                cell.value = val if not isinstance(val, float) else round(val, 1)
            data_style(cell, "center" if col in ("out_stops","airline_count") else "left", bg=bg)

    set_col_widths(ws, {"A":12,"B":9,"C":12,"D":12,"E":10,"F":9,
                        "G":22,"H":10,"I":7,"J":18,"K":18,"L":13})
    freeze(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    return ws


# ── Sheet 2: Price Summary ────────────────────────────────────────────────────

def sheet_summary(wb, df):
    ws = wb.create_sheet("Price Summary")

    summary = df.groupby(["depart_date","season","checked_at"]).agg(
        cheapest=("price_eur","min"),
        average=("price_eur","mean"),
        offers=("price_eur","count"),
        airline_count=("airline_count","max")
    ).reset_index().sort_values(["depart_date","checked_at"])

    summary["cheapest"] = summary["cheapest"].round(0)
    summary["average"]  = summary["average"].round(0)

    headers = ["Depart date","Season","Week checked","Cheapest €","Average €","# Offers","# Airlines"]
    for i, h in enumerate(headers, 1):
        hdr_style(ws.cell(1, i), h)

    for r_idx, (_, row) in enumerate(summary.iterrows(), 2):
        bg = AUG_BG if row["season"] == "August" else DEC_BG
        vals = [
            row["depart_date"].strftime("%Y-%m-%d"),
            row["season"],
            row["checked_at"].strftime("%Y-%m-%d"),
            row["cheapest"],
            row["average"],
            row["offers"],
            row["airline_count"],
        ]
        aligns = ["left","left","left","right","right","center","center"]
        fmts   = [None, None, None, "€#,##0", "€#,##0", None, None]
        for c_idx, (val, aln, fmt) in enumerate(zip(vals, aligns, fmts), 1):
            cell = ws.cell(r_idx, c_idx, val)
            data_style(cell, aln, bg=bg, num_fmt=fmt)

    set_col_widths(ws, {"A":13,"B":9,"C":13,"D":11,"E":11,"F":10,"G":11})
    freeze(ws)
    ws.auto_filter.ref = f"A1:G1"

    # Colour scale on cheapest price column
    if len(summary) > 1:
        last_row = len(summary) + 1
        ws.conditional_formatting.add(f"D2:D{last_row}", ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B"
        ))
    return ws


# ── Sheet 3: Airline Stats ────────────────────────────────────────────────────

def sheet_airlines(wb, df):
    ws = wb.create_sheet("Airline Stats")

    stats = df.groupby("airlines").agg(
        avg_price=("price_eur","mean"),
        min_price=("price_eur","min"),
        max_price=("price_eur","max"),
        times_seen=("price_eur","count")
    ).reset_index().sort_values("avg_price")

    stats["avg_price"] = stats["avg_price"].round(0)
    stats["min_price"] = stats["min_price"].round(0)
    stats["max_price"] = stats["max_price"].round(0)

    headers = ["Airline","Avg price €","Cheapest ever €","Most expensive €","Times seen"]
    for i, h in enumerate(headers, 1):
        hdr_style(ws.cell(1, i), h)

    for r_idx, (_, row) in enumerate(stats.iterrows(), 2):
        bg = ALT if r_idx % 2 == 0 else "FFFFFF"
        vals   = [row["airlines"], row["avg_price"], row["min_price"], row["max_price"], row["times_seen"]]
        aligns = ["left","right","right","right","center"]
        fmts   = [None,"€#,##0","€#,##0","€#,##0",None]
        for c_idx, (val, aln, fmt) in enumerate(zip(vals, aligns, fmts), 1):
            cell = ws.cell(r_idx, c_idx, val)
            data_style(cell, aln, bg=bg, num_fmt=fmt)
        # Bold the cheapest airline
        if r_idx == 2:
            ws.cell(r_idx, 1).font = Font(name="Arial", size=10, bold=True, color="0F6E56")

    set_col_widths(ws, {"A":25,"B":13,"C":16,"D":18,"E":12})
    freeze(ws)
    return ws


# ── Sheet 4: Disruption Log ───────────────────────────────────────────────────

def sheet_disruption(wb, df, no_res):
    ws = wb.create_sheet("Disruption Log")

    ws.cell(1, 1).value = "This sheet tracks two disruption signals:"
    ws.cell(1, 1).font  = Font(name="Arial", size=10, bold=True)
    ws.cell(2, 1).value = "1. Dates with zero flight offers (possible route suspension)"
    ws.cell(3, 1).value = "2. Weeks where the number of airlines dropped significantly"
    ws.cell(2, 1).font  = Font(name="Arial", size=10, italic=True)
    ws.cell(3, 1).font  = Font(name="Arial", size=10, italic=True)

    # Section: No-results log
    ws.cell(5, 1).value = "Zero-results log (dates with NO flights found)"
    ws.cell(5, 1).font  = Font(name="Arial", size=11, bold=True)

    no_res_headers = ["Checked","Depart date","Return date","Note"]
    for i, h in enumerate(no_res_headers, 1):
        hdr_style(ws.cell(6, i), h)

    if len(no_res):
        for r_idx, (_, row) in enumerate(no_res.iterrows(), 7):
            vals = [str(row.get("checked_at","")), str(row.get("depart_date","")),
                    str(row.get("return_date","")), "⚠ No flights returned — possible disruption"]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(r_idx, c_idx, val)
                data_style(cell, bg=WARN)
    else:
        ws.cell(7, 1).value = "No disruptions recorded yet — good sign!"
        ws.cell(7, 1).font  = Font(name="Arial", size=10, color="3B6D11")

    # Section: Airline count drop
    start_row = max(10, 7 + len(no_res) + 2)
    ws.cell(start_row, 1).value = "Airline availability per week (low count = potential disruption)"
    ws.cell(start_row, 1).font  = Font(name="Arial", size=11, bold=True)

    ac_headers = ["Depart date","Season","Week checked","# Airlines","Signal"]
    for i, h in enumerate(ac_headers, 1):
        hdr_style(ws.cell(start_row+1, i), h)

    ac = df.groupby(["depart_date","season","checked_at"])["airline_count"].max().reset_index()
    ac = ac.sort_values(["depart_date","checked_at"])

    for r_idx, (_, row) in enumerate(ac.iterrows(), start_row+2):
        count = int(row["airline_count"])
        signal = "⚠ Low — check news" if count <= 2 else ("✓ Normal" if count >= 4 else "— Monitor")
        bg = WARN if count <= 2 else GOOD if count >= 4 else ALT
        vals = [row["depart_date"].strftime("%Y-%m-%d") if hasattr(row["depart_date"],"strftime") else str(row["depart_date"]),
                row["season"],
                row["checked_at"].strftime("%Y-%m-%d") if hasattr(row["checked_at"],"strftime") else str(row["checked_at"]),
                count, signal]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(r_idx, c_idx, val)
            data_style(cell, "center" if c_idx == 4 else "left", bg=bg)

    set_col_widths(ws, {"A":14,"B":9,"C":14,"D":12,"E":22})
    return ws


# ── Sheet 5: Best Deals ───────────────────────────────────────────────────────

def sheet_best_deals(wb, df):
    ws = wb.create_sheet("Best Deals")

    ws.cell(1, 1).value = "Top 20 cheapest offers ever recorded"
    ws.cell(1, 1).font  = Font(name="Arial", size=12, bold=True)

    best = df.nsmallest(20, "price_eur")
    cols = ["checked_at","season","depart_date","return_date","price_eur","airlines","out_stops","out_dep_time"]
    labels = ["Spotted on","Season","Depart","Return","Price €","Airline","Stops","Dep time"]

    for i, lbl in enumerate(labels, 1):
        hdr_style(ws.cell(3, i), lbl)

    for r_idx, (_, row) in enumerate(best.iterrows(), 4):
        bg = AUG_BG if str(row.get("season","")) == "August" else DEC_BG
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(r_idx, c_idx)
            val  = row[col]
            if pd.isna(val) or val == "":
                cell.value = ""
            elif col in ("checked_at","depart_date","return_date") and hasattr(val,"strftime"):
                cell.value = val.strftime("%Y-%m-%d")
            elif col == "price_eur":
                cell.value = round(float(val), 0)
                data_style(cell, "right", bold=True, bg=bg, num_fmt="€#,##0")
                continue
            else:
                cell.value = val
            data_style(cell, "center" if col == "out_stops" else "left", bg=bg)

    # Gold star on rank 1
    ws.cell(4, 5).font = Font(name="Arial", size=10, bold=True, color="854F0B")
    ws.cell(4, 5).fill = PatternFill("solid", fgColor="FAC775")

    set_col_widths(ws, {"A":13,"B":9,"C":12,"D":12,"E":10,"F":22,"G":7,"H":18})
    freeze(ws, "A4")
    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    df, no_res = load_data()
    if df.empty:
        print("No data to process. Run flight_tracker.py first.")
        return

    print(f"Loaded {len(df)} records. Generating report...")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_raw(wb, df)
    sheet_summary(wb, df)
    sheet_airlines(wb, df)
    sheet_disruption(wb, df, no_res)
    sheet_best_deals(wb, df)

    wb.save(OUT_PATH)
    print(f"Report saved → {OUT_PATH}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
